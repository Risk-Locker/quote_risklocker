"""Shared CSV/Excel import helpers with strict upload validation.

Only CSV and .xlsx are accepted. Excel workbooks are read with openpyxl in
data-only, read-only mode: formulas are never evaluated and macro files
(.xlsm) are rejected by the extension whitelist. All inserts go through the
ORM with parameterized statements, so uploaded content can never execute SQL.
"""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path

from app.core.errors import AppError

MAX_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_ROWS = 5000
MAX_CELL_LEN = 500
_CONTROL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def validate_upload(filename: str, data: bytes) -> None:
    if len(data) > MAX_UPLOAD_BYTES:
        raise AppError("File is too large. Maximum is 5 MB.", 400)
    suffix = Path(filename).suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise AppError("Only .csv or .xlsx files are supported.", 400)


def _clean(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    text = _CONTROL_RE.sub("", text)
    if len(text) > MAX_CELL_LEN:
        text = text[:MAX_CELL_LEN]
    return text or None


def _clean_cell(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    return _clean(value)


def _as_rows(rows: list[list[object]]) -> list[list[object]]:
    if not rows:
        raise AppError("The file contains no data.", 400)
    rows = rows[: MAX_ROWS + 1]
    if len(rows) > MAX_ROWS:
        raise AppError(f"Too many rows. Maximum is {MAX_ROWS}.", 400)
    return rows


def parse_tabular(filename: str, data: bytes) -> list[list[object]]:
    """Parse a CSV or single-sheet XLSX into rows of values."""
    validate_upload(filename, data)
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        try:
            text = data.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = data.decode("latin-1", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = [[_clean_cell(cell) for cell in row] for row in reader if any(str(c).strip() for c in row)]
        return _as_rows(rows)
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:
        raise AppError("Could not read the Excel file. Save it as .xlsx first.", 400) from exc
    ws = wb.active
    rows = [
        [_clean_cell(cell) for cell in row]
        for row in ws.iter_rows(values_only=True)
        if any(c is not None and str(c).strip() for c in row)
    ]
    wb.close()
    return _as_rows(rows)


def parse_vehicles_workbook(filename: str, data: bytes) -> list[tuple[str, list[dict]]]:
    """Parse an XLSX workbook where each sheet is a brand and each column a model.

    Column headers are model names; the cells below a column are search
    terms/aliases for that model. Returns (brand_name, [{name, aliases}]) per sheet.
    CSV input is rejected here; use parse_tabular for the export-format CSV instead.
    """
    validate_upload(filename, data)
    if Path(filename).suffix.lower() != ".xlsx":
        raise AppError("Multi-sheet import requires an .xlsx file.", 400)
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:
        raise AppError("Could not read the Excel file. Save it as .xlsx first.", 400) from exc
    sheets: list[tuple[str, list[dict]]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [
            [_clean_cell(cell) for cell in row]
            for row in ws.iter_rows(values_only=True)
            if any(c is not None and str(c).strip() for c in row)
        ]
        rows = _as_rows(rows)
        if not rows:
            continue
        header = [_clean(cell) for cell in rows[0]]
        alias_lists: dict[int, list[str]] = {}
        for row in rows[1:]:
            for idx, cell in enumerate(row[: len(header)]):
                alias = _clean(cell)
                if alias:
                    alias_lists.setdefault(idx, []).append(alias)
        models = [
            {"name": header[idx], "aliases": alias_lists.get(idx, [])}
            for idx in range(len(header))
            if header[idx]
        ]
        if models:
            sheets.append((sheet_name.strip(), models))
    wb.close()
    if not sheets:
        raise AppError("The workbook contains no data.", 400)
    return sheets
